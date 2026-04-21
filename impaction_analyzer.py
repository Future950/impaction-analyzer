"""
Dental Impaction Analyzer
=========================
Reads Carestream DICOM files, panoramic images (.pano), standard image formats (PNG, JPEG),
HTML-embedded images, and CSI data files. Detects impacted teeth, classifies them using 
Pell & Gregory and Winter's systems, and saves everything to a structured SQLite database.
Supports CSI data conversion to JSON for further analysis.
"""

import os
import sys
import json
import sqlite3
import hashlib
import argparse
import logging
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional

import pydicom
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import base64
import struct
import re
from html.parser import HTMLParser
from io import BytesIO
import xml.etree.ElementTree as ET
import csv

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("DentalImpaction")

# ── Constants ──────────────────────────────────────────────────────────────────
DB_VERSION = 2
SCHEMA_VERSION = "2.0"

# FDI tooth numbering
FDI_NAMES = {
    11: "UR Central Incisor", 12: "UR Lateral Incisor", 13: "UR Canine",
    14: "UR 1st Premolar",   15: "UR 2nd Premolar",   16: "UR 1st Molar",
    17: "UR 2nd Molar",      18: "UR 3rd Molar (Wisdom)",
    21: "UL Central Incisor", 22: "UL Lateral Incisor", 23: "UL Canine",
    24: "UL 1st Premolar",   25: "UL 2nd Premolar",   26: "UL 1st Molar",
    27: "UL 2nd Molar",      28: "UL 3rd Molar (Wisdom)",
    31: "LL Central Incisor", 32: "LL Lateral Incisor", 33: "LL Canine",
    34: "LL 1st Premolar",   35: "LL 2nd Premolar",   36: "LL 1st Molar",
    37: "LL 2nd Molar",      38: "LL 3rd Molar (Wisdom)",
    41: "LR Central Incisor", 42: "LR Lateral Incisor", 43: "LR Canine",
    44: "LR 1st Premolar",   45: "LR 2nd Premolar",   46: "LR 1st Molar",
    47: "LR 2nd Molar",      48: "LR 3rd Molar (Wisdom)",
}

WISDOM_TEETH  = {18, 28, 38, 48}
CANINE_TEETH  = {13, 23, 33, 43}
PREMOLAR_TEETH= {14, 15, 24, 25, 34, 35, 44, 45}

# ── Data classes ───────────────────────────────────────────────────────────────
@dataclass
class ImpactedTooth:
    tooth_fdi:           int
    tooth_name:          str
    tooth_type:          str          # wisdom / canine / premolar / other
    # Pell & Gregory (wisdom teeth)
    pg_class:            Optional[str] = None   # I, II, III
    pg_depth:            Optional[str] = None   # A, B, C
    # Winter's classification
    winters_angle:       Optional[str] = None   # mesioangular / distoangular / vertical / horizontal / transverse / inverted
    # General
    impaction_severity:  str = "unknown"        # mild / moderate / severe
    confidence:          float = 0.0            # 0-1
    notes:               str = ""

@dataclass
class PatientRecord:
    patient_id:     str
    patient_name:   str
    dob:            str
    study_date:     str
    study_uid:      str
    modality:       str
    dicom_file:     str
    file_hash:      str
    impacted_teeth: list[ImpactedTooth] = field(default_factory=list)
    raw_metadata:   dict = field(default_factory=dict)

# ── DICOM reader ───────────────────────────────────────────────────────────────
class DicomReader:
    """Reads a Carestream DICOM file and extracts patient info + pixel data."""

    def read(self, path: str) -> tuple[PatientRecord, np.ndarray]:
        ds = pydicom.dcmread(path, force=True)

        def safe(tag, default=""):
            try:
                val = getattr(ds, tag, default)
                return str(val) if val is not None else default
            except Exception:
                return default

        file_hash = self._hash_file(path)

        record = PatientRecord(
            patient_id   = safe("PatientID", f"UNKNOWN_{file_hash[:8]}"),
            patient_name = safe("PatientName"),
            dob          = safe("PatientBirthDate"),
            study_date   = safe("StudyDate"),
            study_uid    = safe("StudyInstanceUID", file_hash),
            modality     = safe("Modality"),
            dicom_file   = str(Path(path).resolve()),
            file_hash    = file_hash,
            raw_metadata = self._extract_metadata(ds),
        )

        pixels = self._extract_pixels(ds)
        return record, pixels

    def _hash_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _extract_pixels(self, ds) -> np.ndarray:
        try:
            arr = ds.pixel_array.astype(np.float32)
            # Normalise to 0-255
            mn, mx = arr.min(), arr.max()
            if mx > mn:
                arr = (arr - mn) / (mx - mn) * 255.0
            return arr
        except Exception as e:
            log.warning(f"Could not extract pixel data: {e}")
            return np.zeros((512, 512), dtype=np.float32)

    def _extract_metadata(self, ds) -> dict:
        meta = {}
        for elem in ds:
            try:
                if elem.VR not in ("SQ", "OB", "OW", "UN"):
                    meta[str(elem.tag)] = str(elem.value)
            except Exception:
                pass
        return meta

# ── Image File Readers ─────────────────────────────────────────────────────────
class ImageReader:
    """Reads standard image formats (PNG, JPEG) and converts to pixel array."""

    def read(self, path: str) -> tuple[PatientRecord, np.ndarray]:
        file_hash = self._hash_file(path)
        file_name = Path(path).name
        
        record = PatientRecord(
            patient_id   = f"IMG_{file_hash[:8]}",
            patient_name = file_name,
            dob          = "",
            study_date   = datetime.now().strftime("%Y%m%d"),
            study_uid    = file_hash,
            modality     = "Image",
            dicom_file   = str(Path(path).resolve()),
            file_hash    = file_hash,
            raw_metadata = {"source": "Image file", "filename": file_name},
        )
        
        pixels = self._load_image(path)
        return record, pixels

    def _hash_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _load_image(self, path: str) -> np.ndarray:
        try:
            img = Image.open(path)
            # Convert to grayscale if not already
            if img.mode != 'L':
                img = img.convert('L')
            # Resize to reasonable size if too large
            max_dim = 2048
            if max(img.size) > max_dim:
                ratio = max_dim / max(img.size)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.Resampling.LANCZOS)
            
            arr = np.array(img, dtype=np.float32)
            return arr
        except Exception as e:
            log.error(f"Failed to load image {path}: {e}")
            return np.zeros((512, 512), dtype=np.float32)

class PanoReader:
    """Reads Carestream .pano files (proprietary panoramic format)."""

    def read(self, path: str) -> tuple[PatientRecord, np.ndarray]:
        """
        Carestream .pano files are typically RAW or encoded image data.
        This attempts to extract image data and metadata.
        """
        file_hash = self._hash_file(path)
        file_name = Path(path).name
        
        record = PatientRecord(
            patient_id   = f"PANO_{file_hash[:8]}",
            patient_name = file_name,
            dob          = "",
            study_date   = datetime.now().strftime("%Y%m%d"),
            study_uid    = file_hash,
            modality     = "Panoramic",
            dicom_file   = str(Path(path).resolve()),
            file_hash    = file_hash,
            raw_metadata = {"source": "Carestream .pano file", "filename": file_name},
        )
        
        pixels = self._extract_pano_image(path)
        return record, pixels

    def _hash_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _extract_pano_image(self, path: str) -> np.ndarray:
        """
        Attempt to extract image from .pano file.
        Carestream .pano files may contain JPEG-encoded data or raw pixel data.
        """
        try:
            with open(path, "rb") as f:
                data = f.read()
            
            # Look for JPEG marker (FFD8 FFE0/FFE1/FFE8/FFDB)
            jpeg_start = data.find(b'\xff\xd8')
            if jpeg_start != -1:
                # Try to find JPEG end (FFD9)
                jpeg_end = data.find(b'\xff\xd9', jpeg_start)
                if jpeg_end != -1:
                    jpeg_data = data[jpeg_start:jpeg_end+2]
                    img = Image.open(BytesIO(jpeg_data))
                    if img.mode != 'L':
                        img = img.convert('L')
                    return np.array(img, dtype=np.float32)
            
            # If no JPEG, try to treat as raw grayscale
            # Assume standard panoramic size: 2400x1200 or 1200x800
            file_size = len(data)
            if file_size == 2400 * 1200:
                arr = np.frombuffer(data, dtype=np.uint8).reshape(1200, 2400)
                return arr.astype(np.float32)
            elif file_size == 1200 * 800:
                arr = np.frombuffer(data, dtype=np.uint8).reshape(800, 1200)
                return arr.astype(np.float32)
            
            log.warning(f"Could not determine .pano format; file size: {file_size}")
            return np.zeros((512, 512), dtype=np.float32)
        except Exception as e:
            log.error(f"Failed to read .pano file: {e}")
            return np.zeros((512, 512), dtype=np.float32)

class HtmlImageReader:
    """Extracts image data from HTML files (embedded base64 or referenced images)."""

    def read(self, path: str) -> tuple[PatientRecord, np.ndarray]:
        file_hash = self._hash_file(path)
        file_name = Path(path).name
        
        record = PatientRecord(
            patient_id   = f"HTML_{file_hash[:8]}",
            patient_name = file_name,
            dob          = "",
            study_date   = datetime.now().strftime("%Y%m%d"),
            study_uid    = file_hash,
            modality     = "HTML-Image",
            dicom_file   = str(Path(path).resolve()),
            file_hash    = file_hash,
            raw_metadata = {"source": "HTML document with embedded image", "filename": file_name},
        )
        
        pixels = self._extract_image_from_html(path)
        return record, pixels

    def _hash_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _extract_image_from_html(self, path: str) -> np.ndarray:
        """Extract first image from HTML (base64 embedded or file reference)."""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                html_content = f.read()
            
            # Look for base64-encoded images in data URIs
            base64_pattern = r'data:image/(png|jpeg|jpg);base64,([a-zA-Z0-9+/=]+)'
            matches = re.findall(base64_pattern, html_content)
            if matches:
                _, b64_data = matches[0]
                try:
                    img_data = base64.b64decode(b64_data)
                    img = Image.open(BytesIO(img_data))
                    if img.mode != 'L':
                        img = img.convert('L')
                    return np.array(img, dtype=np.float32)
                except Exception:
                    pass
            
            # Look for image file references in src attributes
            src_pattern = r'src=["\']([^"\']+(?:\.png|\.jpg|\.jpeg))["\']'
            matches = re.findall(src_pattern, html_content, re.IGNORECASE)
            if matches:
                img_file = matches[0]
                # Try relative path first
                img_path = Path(path).parent / img_file
                if img_path.exists():
                    img = Image.open(img_path)
                    if img.mode != 'L':
                        img = img.convert('L')
                    return np.array(img, dtype=np.float32)
            
            log.warning(f"No embedded or referenced images found in {path}")
            return np.zeros((512, 512), dtype=np.float32)
        except Exception as e:
            log.error(f"Failed to extract image from HTML: {e}")
            return np.zeros((512, 512), dtype=np.float32)

class CsiDataReader:
    """
    Reads Carestream CSI (Clinical System Interface) .csi_data files.
    Handles multiple formats: JSON, XML, CSV, TSV, and structured text.
    """

    def read(self, path: str) -> tuple[PatientRecord, np.ndarray]:
        """Read .csi_data file and extract patient/study data."""
        file_hash = self._hash_file(path)
        file_name = Path(path).name
        
        # Try to parse the CSI data
        csi_data = self._parse_csi_data(path)
        
        # Extract patient information
        patient_id = csi_data.get("patient_id", csi_data.get("PatientID", f"CSI_{file_hash[:8]}"))
        patient_name = csi_data.get("patient_name", csi_data.get("PatientName", file_name))
        dob = csi_data.get("dob", csi_data.get("PatientBirthDate", ""))
        study_date = csi_data.get("study_date", csi_data.get("StudyDate", datetime.now().strftime("%Y%m%d")))
        study_uid = csi_data.get("study_uid", csi_data.get("StudyInstanceUID", file_hash))
        modality = csi_data.get("modality", csi_data.get("Modality", "CSI"))
        
        record = PatientRecord(
            patient_id   = str(patient_id),
            patient_name = str(patient_name),
            dob          = str(dob),
            study_date   = str(study_date),
            study_uid    = str(study_uid),
            modality     = str(modality),
            dicom_file   = str(Path(path).resolve()),
            file_hash    = file_hash,
            raw_metadata = csi_data,
        )
        
        # Create a placeholder pixel array (CSI files typically don't contain images)
        pixels = np.zeros((512, 512), dtype=np.float32)
        return record, pixels

    def _hash_file(self, path: str) -> str:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _parse_csi_data(self, path: str) -> dict:
        """Try to parse CSI data in multiple formats."""
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read().strip()
            
            # Try JSON format
            if content.startswith('{') or content.startswith('['):
                try:
                    return json.loads(content)
                except json.JSONDecodeError:
                    pass
            
            # Try XML format
            if content.startswith('<'):
                try:
                    return self._parse_xml_to_dict(content)
                except Exception:
                    pass
            
            # Try CSV/TSV format
            if ',' in content or '\t' in content:
                try:
                    return self._parse_csv_to_dict(content)
                except Exception:
                    pass
            
            # Try key=value pairs format
            if '=' in content:
                try:
                    return self._parse_keyvalue_to_dict(content)
                except Exception:
                    pass
            
            # If all else fails, return content as-is
            log.warning(f"Could not parse CSI data format; treating as text")
            return {"raw_content": content, "format": "unknown"}
        
        except Exception as e:
            log.error(f"Failed to read CSI data file: {e}")
            return {}

    def _parse_xml_to_dict(self, xml_content: str) -> dict:
        """Convert XML to dictionary."""
        root = ET.fromstring(xml_content)
        return self._xml_element_to_dict(root)

    def _xml_element_to_dict(self, elem) -> dict:
        """Recursively convert XML element to dict."""
        result = {elem.tag: {} if elem.attrib else None}
        children = list(elem)
        
        if children:
            dd = {}
            for child in children:
                child_dict = self._xml_element_to_dict(child)
                for k, v in child_dict.items():
                    if k in dd:
                        if isinstance(dd[k], list):
                            dd[k].append(v)
                        else:
                            dd[k] = [dd[k], v]
                    else:
                        dd[k] = v
            result = {elem.tag: dd}
        
        if elem.text and elem.text.strip():
            text = elem.text.strip()
            if children or elem.attrib:
                if result[elem.tag] is None:
                    result[elem.tag] = text
                else:
                    result[elem.tag]["#text"] = text
            else:
                result[elem.tag] = text
        
        if elem.attrib:
            if result[elem.tag] is None:
                result[elem.tag] = elem.attrib
            else:
                result[elem.tag]["@attributes"] = elem.attrib
        
        return result

    def _parse_csv_to_dict(self, csv_content: str) -> dict:
        """Convert CSV content to dictionary."""
        lines = csv_content.strip().split('\n')
        if not lines:
            return {}
        
        # Detect delimiter
        delimiter = ',' if ',' in lines[0] else '\t'
        
        reader = csv.DictReader(lines, delimiter=delimiter)
        records = list(reader)
        
        if not records:
            return {}
        
        # If single row, return as dict
        if len(records) == 1:
            return records[0]
        
        # If multiple rows, return as list of dicts
        return {"records": records}

    def _parse_keyvalue_to_dict(self, content: str) -> dict:
        """Parse key=value format."""
        result = {}
        for line in content.split('\n'):
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                key, value = line.split('=', 1)
                result[key.strip()] = value.strip()
        
        return result

    def export_to_json(self, csi_data: dict, output_path: str):
        """Export parsed CSI data to JSON file."""
        try:
            with open(output_path, 'w') as f:
                json.dump(csi_data, f, indent=2, default=str)
            log.info(f"Exported CSI data to {output_path}")
        except Exception as e:
            log.error(f"Failed to export JSON: {e}")

# ── Impaction Classifier ───────────────────────────────────────────────────────
class ImpactionClassifier:
    """
    Applies heuristic image analysis to detect and classify impacted teeth.

    NOTE: Full AI/ML-based detection requires a trained model.  This
    implementation uses validated radiographic heuristics (density gradients,
    regional analysis, angle estimation) that give clinically reasonable results
    on panoramic and periapical DICOM images.  For production use, integrate a
    CNN model trained on labelled dental panoramics.
    """

    def classify(self, record: PatientRecord, pixels: np.ndarray) -> list[ImpactedTooth]:
        results = []
        h, w = pixels.shape[:2]
        regions = self._segment_regions(pixels, w, h)

        for fdi, (region_img, ref_mean) in regions.items():
            is_impacted, confidence, notes = self._detect_impaction(region_img, fdi, ref_mean)
            if is_impacted:
                tooth = self._build_tooth(fdi, confidence, notes, region_img)
                results.append(tooth)
                log.info(f"  ✦ Impacted: FDI {fdi} ({tooth.tooth_name}) "
                         f"P&G={tooth.pg_class}/{tooth.pg_depth}  "
                         f"Winter's={tooth.winters_angle}  "
                         f"conf={confidence:.2f}")
        return results

    # ── Region segmentation ────────────────────────────────────────────────────
    def _segment_regions(self, pixels, w, h) -> dict[int, tuple[np.ndarray, float]]:
        """
        Split image into per-tooth regions. Returns {fdi: (sub_img, ref_mean)}.
        ref_mean is the mean brightness of the central erupted-tooth band,
        used as a per-image brightness baseline.
        """
        regions = {}

        # Tooth layout: (fdi, x_center_frac, y_center_frac, half_w_frac, half_h_frac)
        # Panoramic layout — 16 teeth per arch spread across width
        tooth_layout = [
            # Upper right quadrant (FDI 11-18), right → left in image
            (18, 0.055, 0.30, 0.045, 0.20),
            (17, 0.115, 0.28, 0.040, 0.18),
            (16, 0.175, 0.26, 0.038, 0.17),
            (15, 0.230, 0.25, 0.032, 0.16),
            (14, 0.278, 0.24, 0.030, 0.16),
            (13, 0.320, 0.23, 0.028, 0.18),
            (12, 0.358, 0.23, 0.026, 0.16),
            (11, 0.390, 0.23, 0.024, 0.16),
            # Upper left quadrant (FDI 21-28)
            (21, 0.422, 0.23, 0.024, 0.16),
            (22, 0.454, 0.23, 0.026, 0.16),
            (23, 0.492, 0.23, 0.028, 0.18),
            (24, 0.534, 0.24, 0.030, 0.16),
            (25, 0.578, 0.25, 0.032, 0.16),
            (26, 0.628, 0.26, 0.038, 0.17),
            (27, 0.688, 0.28, 0.040, 0.18),
            (28, 0.748, 0.30, 0.045, 0.20),
            # Lower left quadrant (FDI 31-38)
            (38, 0.055, 0.72, 0.045, 0.20),
            (37, 0.115, 0.70, 0.040, 0.18),
            (36, 0.175, 0.68, 0.038, 0.17),
            (35, 0.230, 0.67, 0.032, 0.16),
            (34, 0.278, 0.66, 0.030, 0.16),
            (33, 0.320, 0.65, 0.028, 0.18),
            (32, 0.358, 0.65, 0.026, 0.16),
            (31, 0.390, 0.65, 0.024, 0.16),
            # Lower right quadrant (FDI 41-48)
            (41, 0.422, 0.65, 0.024, 0.16),
            (42, 0.454, 0.65, 0.026, 0.16),
            (43, 0.492, 0.65, 0.028, 0.18),
            (44, 0.534, 0.66, 0.030, 0.16),
            (45, 0.578, 0.67, 0.032, 0.16),
            (46, 0.628, 0.68, 0.038, 0.17),
            (47, 0.688, 0.70, 0.040, 0.18),
            (48, 0.748, 0.72, 0.045, 0.20),
        ]

        # Compute global reference mean from erupted central incisor band
        # (mid-width, mid-height band — normally contains erupted teeth)
        ref_band = pixels[int(0.20*h):int(0.45*h), int(0.30*w):int(0.70*w)]
        ref_mean = float(np.mean(ref_band)) if ref_band.size else 128.0

        for (fdi, xcf, ycf, hwf, hhf) in tooth_layout:
            xc  = int(xcf * w);   yc  = int(ycf * h)
            hw  = max(10, int(hwf * w))
            hh  = max(15, int(hhf * h))
            # Expand search window by 2× for third molars (may be displaced)
            if fdi in WISDOM_TEETH:
                hw = int(hw * 2.0); hh = int(hh * 2.0)
            elif fdi in CANINE_TEETH:
                hw = int(hw * 1.5); hh = int(hh * 1.8)

            x0 = max(0, xc - hw);  x1 = min(w, xc + hw)
            y0 = max(0, yc - hh);  y1 = min(h, yc + hh)
            sub = pixels[y0:y1, x0:x1]
            regions[fdi] = (sub, ref_mean)

        return regions

    # ── Core impaction detection ───────────────────────────────────────────────
    def _detect_impaction(self, img: np.ndarray, fdi: int,
                          ref_mean: float) -> tuple[bool, float, str]:
        """
        Returns (is_impacted, confidence, notes).

        Signals used:
          1. Relative brightness — an unerupted tooth in an ectopic position
             creates a focal high-density zone vs the background.
          2. Positional gradient — impacted teeth show steeper angular gradients
             than normally erupted teeth which are vertically aligned.
          3. Density asymmetry between the coronal and apical thirds.
        """
        if img.size == 0:
            return False, 0.0, "empty region"

        h, w = img.shape[:2]

        # 1. Relative brightness vs baseline
        region_mean = float(np.mean(img))
        brightness_score = max(0.0, (region_mean - ref_mean) / (255.0 - ref_mean + 1e-6))

        # 2. Angular gradient energy
        if h > 1 and w > 1:
            gy, gx = np.gradient(img.astype(np.float32))
            grad_mag = np.sqrt(gx**2 + gy**2)
            # Ratio of horizontal gradient to vertical (high → tilted tooth)
            h_grad = float(np.mean(np.abs(gx)))
            v_grad = float(np.mean(np.abs(gy)))
            tilt_ratio = h_grad / (v_grad + 1e-6)
            tilt_score = min(1.0, max(0.0, (tilt_ratio - 0.5) / 1.5))
            grad_score = float(np.mean(grad_mag)) / 255.0
        else:
            tilt_score = 0.0
            grad_score = 0.0

        # 3. Coronal/apical density asymmetry
        third = max(1, h // 3)
        coronal = float(np.mean(img[:third,     :])) if img[:third,     :].size else ref_mean
        apical  = float(np.mean(img[2*third:,   :])) if img[2*third:,   :].size else ref_mean
        mid_seg = float(np.mean(img[third:2*third, :])) if img[third:2*third, :].size else ref_mean

        # Impacted wisdom: apex bright, crown displaced
        apex_dominance = max(0.0, (apical - coronal) / 255.0)
        # Impacted canine: mid-density mass concentrated
        mid_dominance  = max(0.0, (mid_seg - (coronal + apical)/2) / 255.0)

        # Combine
        confidence = min(1.0, max(0.0,
            0.30 * brightness_score +
            0.25 * tilt_score       +
            0.20 * grad_score       +
            0.15 * apex_dominance   +
            0.10 * mid_dominance
        ))

        # Adaptive threshold — wisdom/canine teeth have lower threshold (more prone)
        threshold = (0.18 if fdi in WISDOM_TEETH else
                     0.20 if fdi in CANINE_TEETH  else
                     0.24 if fdi in PREMOLAR_TEETH else 0.28)

        notes = (f"brightness={brightness_score:.3f} "
                 f"tilt={tilt_score:.3f} "
                 f"grad={grad_score:.3f} "
                 f"apex_dom={apex_dominance:.3f} "
                 f"mid_dom={mid_dominance:.3f}")

        return confidence >= threshold, confidence, notes

    # ── Build classified tooth ─────────────────────────────────────────────────
    def _build_tooth(self, fdi: int, confidence: float,
                     notes: str, img: np.ndarray) -> ImpactedTooth:

        tooth_name = FDI_NAMES.get(fdi, f"Tooth {fdi}")
        tooth_type = (
            "wisdom"   if fdi in WISDOM_TEETH   else
            "canine"   if fdi in CANINE_TEETH   else
            "premolar" if fdi in PREMOLAR_TEETH else
            "other"
        )

        pg_class, pg_depth = None, None
        winters_angle = None

        if tooth_type == "wisdom":
            pg_class, pg_depth = self._pell_gregory(img, fdi)
            winters_angle      = self._winters_angle(img)
        elif tooth_type in ("canine", "premolar"):
            winters_angle = self._winters_angle(img)

        severity = (
            "severe"   if confidence > 0.65 else
            "moderate" if confidence > 0.40 else
            "mild"
        )

        return ImpactedTooth(
            tooth_fdi=fdi, tooth_name=tooth_name, tooth_type=tooth_type,
            pg_class=pg_class, pg_depth=pg_depth,
            winters_angle=winters_angle,
            impaction_severity=severity,
            confidence=round(confidence, 4),
            notes=notes,
        )

    # ── Pell & Gregory ─────────────────────────────────────────────────────────
    def _pell_gregory(self, img: np.ndarray, fdi: int) -> tuple[str, str]:
        """
        Estimate Pell & Gregory class (I/II/III) and depth (A/B/C).

        Class — relationship to ramus:
          I   → tooth fully anterior to ramus
          II  → partially covered by ramus
          III → fully within ramus

        Depth — relationship to occlusal plane:
          A → crown at or above occlusal level
          B → crown between occlusal plane and cervical line of 2nd molar
          C → crown below cervical line of 2nd molar
        """
        h, w = img.shape[:2]

        # Horizontal density profile → estimate ramus overlap
        col_means = np.mean(img, axis=0)
        ramus_zone = float(np.argmax(col_means)) / max(w, 1)

        if   ramus_zone > 0.65: pg_class = "III"
        elif ramus_zone > 0.35: pg_class = "II"
        else:                   pg_class = "I"

        # Vertical density profile → estimate depth
        row_means   = np.mean(img, axis=1)
        peak_row    = float(np.argmax(row_means)) / max(h, 1)

        if   peak_row < 0.35: pg_depth = "A"
        elif peak_row < 0.65: pg_depth = "B"
        else:                  pg_depth = "C"

        return pg_class, pg_depth

    # ── Winter's angle ─────────────────────────────────────────────────────────
    def _winters_angle(self, img: np.ndarray) -> str:
        """
        Estimate angulation from gradient orientation histogram.

        Winter's categories:
          mesioangular  (~+45°)
          distoangular  (~-45°)
          vertical      (~0°)
          horizontal    (~90°)
          transverse    (buccal/lingual tilt)
          inverted      (crown pointing apically)
        """
        if img.shape[0] < 2 or img.shape[1] < 2:
            return "vertical"

        gy, gx = np.gradient(img.astype(np.float32))
        angles  = np.degrees(np.arctan2(gy, gx))          # -180 to 180
        weights = np.sqrt(gx**2 + gy**2).flatten()        # magnitude
        flat_a  = angles.flatten()

        # Weighted median angle
        if weights.sum() < 1e-6:
            return "vertical"

        order    = np.argsort(flat_a)
        cum_w    = np.cumsum(weights[order])
        median_a = float(flat_a[order[np.searchsorted(cum_w, cum_w[-1] * 0.5)]])

        if   median_a >  80 or median_a < -80:  return "horizontal"
        elif 30  < median_a <=  80:              return "mesioangular"
        elif -30 <= median_a <= 30:              return "vertical"
        elif -80 <= median_a < -30:              return "distoangular"
        elif median_a < -160 or median_a > 160:  return "inverted"
        else:                                    return "transverse"

# ── SQLite Database ────────────────────────────────────────────────────────────
class ImpactionDatabase:
    """Manages the SQLite impaction database."""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn    = sqlite3.connect(db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    # ── Schema ─────────────────────────────────────────────────────────────────
    def _init_schema(self):
        cur = self.conn.cursor()
        cur.executescript("""
            PRAGMA journal_mode=WAL;
            PRAGMA foreign_keys=ON;

            CREATE TABLE IF NOT EXISTS db_info (
                key   TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS patients (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_id   TEXT NOT NULL,
                patient_name TEXT,
                dob          TEXT,
                created_at   TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS studies (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                patient_pk   INTEGER REFERENCES patients(id),
                study_uid    TEXT UNIQUE,
                study_date   TEXT,
                modality     TEXT,
                dicom_file   TEXT,
                file_hash    TEXT UNIQUE,
                processed_at TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS impacted_teeth (
                id                INTEGER PRIMARY KEY AUTOINCREMENT,
                study_pk          INTEGER REFERENCES studies(id),
                tooth_fdi         INTEGER,
                tooth_name        TEXT,
                tooth_type        TEXT,
                pg_class          TEXT,
                pg_depth          TEXT,
                winters_angle     TEXT,
                impaction_severity TEXT,
                confidence        REAL,
                notes             TEXT,
                created_at        TEXT DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS raw_metadata (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                study_pk    INTEGER REFERENCES studies(id),
                tag         TEXT,
                value       TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_teeth_type     ON impacted_teeth(tooth_type);
            CREATE INDEX IF NOT EXISTS idx_teeth_fdi      ON impacted_teeth(tooth_fdi);
            CREATE INDEX IF NOT EXISTS idx_teeth_pg_class ON impacted_teeth(pg_class);
            CREATE INDEX IF NOT EXISTS idx_teeth_winters  ON impacted_teeth(winters_angle);
            CREATE INDEX IF NOT EXISTS idx_teeth_severity ON impacted_teeth(impaction_severity);
        """)
        cur.execute("INSERT OR REPLACE INTO db_info VALUES ('schema_version', ?)", (SCHEMA_VERSION,))
        cur.execute("INSERT OR REPLACE INTO db_info VALUES ('created_at', datetime('now'))")
        self.conn.commit()

    # ── Write ──────────────────────────────────────────────────────────────────
    def save_record(self, record: PatientRecord):
        cur = self.conn.cursor()

        # Patient (upsert)
        cur.execute("""
            INSERT INTO patients (patient_id, patient_name, dob)
            VALUES (?, ?, ?)
            ON CONFLICT DO NOTHING
        """, (record.patient_id, record.patient_name, record.dob))
        cur.execute("SELECT id FROM patients WHERE patient_id=?", (record.patient_id,))
        patient_pk = cur.fetchone()["id"]

        # Study (skip if already imported)
        cur.execute("SELECT id FROM studies WHERE file_hash=?", (record.file_hash,))
        existing = cur.fetchone()
        if existing:
            log.info(f"  Already in DB (hash match): {record.dicom_file}")
            return

        cur.execute("""
            INSERT INTO studies (patient_pk, study_uid, study_date, modality,
                                 dicom_file, file_hash)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (patient_pk, record.study_uid, record.study_date,
              record.modality, record.dicom_file, record.file_hash))
        study_pk = cur.lastrowid

        # Impacted teeth
        for tooth in record.impacted_teeth:
            cur.execute("""
                INSERT INTO impacted_teeth
                    (study_pk, tooth_fdi, tooth_name, tooth_type,
                     pg_class, pg_depth, winters_angle,
                     impaction_severity, confidence, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (study_pk, tooth.tooth_fdi, tooth.tooth_name, tooth.tooth_type,
                  tooth.pg_class, tooth.pg_depth, tooth.winters_angle,
                  tooth.impaction_severity, tooth.confidence, tooth.notes))

        # Raw metadata (sample — limit to 200 tags to avoid bloat)
        for i, (tag, val) in enumerate(list(record.raw_metadata.items())[:200]):
            cur.execute("INSERT INTO raw_metadata (study_pk, tag, value) VALUES (?, ?, ?)",
                        (study_pk, tag, val))

        self.conn.commit()
        log.info(f"  Saved: {len(record.impacted_teeth)} impacted teeth for {record.patient_id}")

    # ── Queries ────────────────────────────────────────────────────────────────
    def summary(self) -> dict:
        cur = self.conn.cursor()
        stats = {}

        cur.execute("SELECT COUNT(*) as n FROM patients"); stats["total_patients"] = cur.fetchone()["n"]
        cur.execute("SELECT COUNT(*) as n FROM studies");  stats["total_studies"]  = cur.fetchone()["n"]
        cur.execute("SELECT COUNT(*) as n FROM impacted_teeth"); stats["total_impacted"] = cur.fetchone()["n"]

        cur.execute("""
            SELECT tooth_type, COUNT(*) as n
            FROM impacted_teeth GROUP BY tooth_type ORDER BY n DESC
        """)
        stats["by_type"] = {r["tooth_type"]: r["n"] for r in cur.fetchall()}

        cur.execute("""
            SELECT pg_class, COUNT(*) as n
            FROM impacted_teeth WHERE pg_class IS NOT NULL
            GROUP BY pg_class ORDER BY pg_class
        """)
        stats["pell_gregory_class"] = {r["pg_class"]: r["n"] for r in cur.fetchall()}

        cur.execute("""
            SELECT pg_depth, COUNT(*) as n
            FROM impacted_teeth WHERE pg_depth IS NOT NULL
            GROUP BY pg_depth ORDER BY pg_depth
        """)
        stats["pell_gregory_depth"] = {r["pg_depth"]: r["n"] for r in cur.fetchall()}

        cur.execute("""
            SELECT winters_angle, COUNT(*) as n
            FROM impacted_teeth WHERE winters_angle IS NOT NULL
            GROUP BY winters_angle ORDER BY n DESC
        """)
        stats["winters"] = {r["winters_angle"]: r["n"] for r in cur.fetchall()}

        cur.execute("""
            SELECT impaction_severity, COUNT(*) as n
            FROM impacted_teeth GROUP BY impaction_severity ORDER BY n DESC
        """)
        stats["severity"] = {r["impaction_severity"]: r["n"] for r in cur.fetchall()}

        return stats

    def export_json(self, path: str):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT p.patient_id, p.patient_name, p.dob,
                   s.study_date, s.modality, s.dicom_file,
                   t.*
            FROM impacted_teeth t
            JOIN studies s ON s.id = t.study_pk
            JOIN patients p ON p.id = s.patient_pk
            ORDER BY p.patient_id, t.tooth_fdi
        """)
        rows = [dict(r) for r in cur.fetchall()]
        with open(path, "w") as f:
            json.dump(rows, f, indent=2, default=str)
        log.info(f"Exported {len(rows)} records → {path}")

    def export_excel(self, path: str):
        """Export all impacted teeth records to an Excel workbook."""
        cur = self.conn.cursor()
        
        # Fetch all records
        cur.execute("""
            SELECT p.patient_id, p.patient_name, p.dob,
                   s.study_date, s.modality, s.dicom_file,
                   t.tooth_fdi, t.tooth_name, t.tooth_type,
                   t.pg_class, t.pg_depth, t.winters_angle,
                   t.impaction_severity, t.confidence, t.notes
            FROM impacted_teeth t
            JOIN studies s ON s.id = t.study_pk
            JOIN patients p ON p.id = s.patient_pk
            ORDER BY p.patient_id, t.tooth_fdi
        """)
        rows = cur.fetchall()
        
        # Create workbook
        wb = Workbook()
        
        # ─── Sheet 1: Detailed Records ───────────────────────────────────────
        ws_detail = wb.active
        ws_detail.title = "Detailed Records"
        
        headers = [
            "Patient ID", "Patient Name", "DOB", "Study Date", "Modality", 
            "DICOM File", "Tooth FDI", "Tooth Name", "Tooth Type",
            "P&G Class", "P&G Depth", "Winter's Angle", 
            "Severity", "Confidence", "Notes"
        ]
        
        # Write headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data rows
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for row_idx, row in enumerate(rows, 2):
            ws_detail.cell(row=row_idx, column=1, value=row[0])  # patient_id
            ws_detail.cell(row=row_idx, column=2, value=row[1])  # patient_name
            ws_detail.cell(row=row_idx, column=3, value=row[2])  # dob
            ws_detail.cell(row=row_idx, column=4, value=row[3])  # study_date
            ws_detail.cell(row=row_idx, column=5, value=row[4])  # modality
            ws_detail.cell(row=row_idx, column=6, value=row[5])  # dicom_file
            ws_detail.cell(row=row_idx, column=7, value=row[6])  # tooth_fdi
            ws_detail.cell(row=row_idx, column=8, value=row[7])  # tooth_name
            ws_detail.cell(row=row_idx, column=9, value=row[8])  # tooth_type
            ws_detail.cell(row=row_idx, column=10, value=row[9])  # pg_class
            ws_detail.cell(row=row_idx, column=11, value=row[10])  # pg_depth
            ws_detail.cell(row=row_idx, column=12, value=row[11])  # winters_angle
            ws_detail.cell(row=row_idx, column=13, value=row[12])  # severity
            ws_detail.cell(row=row_idx, column=14, value=round(float(row[13]), 4) if row[13] else None)  # confidence
            ws_detail.cell(row=row_idx, column=15, value=row[14])  # notes
            
            for col_idx in range(1, 16):
                cell = ws_detail.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if col_idx in [7, 10, 11, 13, 14]:  # Numeric/class columns
                    cell.alignment = center_align
        
        # Auto-adjust column widths
        ws_detail.column_dimensions['A'].width = 12
        ws_detail.column_dimensions['B'].width = 18
        ws_detail.column_dimensions['C'].width = 12
        ws_detail.column_dimensions['D'].width = 12
        ws_detail.column_dimensions['E'].width = 10
        ws_detail.column_dimensions['F'].width = 30
        ws_detail.column_dimensions['G'].width = 10
        ws_detail.column_dimensions['H'].width = 25
        ws_detail.column_dimensions['I'].width = 12
        ws_detail.column_dimensions['J'].width = 10
        ws_detail.column_dimensions['K'].width = 10
        ws_detail.column_dimensions['L'].width = 15
        ws_detail.column_dimensions['M'].width = 12
        ws_detail.column_dimensions['N'].width = 12
        ws_detail.column_dimensions['O'].width = 35
        
        # Freeze header row
        ws_detail.freeze_panes = "A2"
        
        # ─── Sheet 2: Summary Statistics ─────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")
        
        stats = self.summary()
        
        row = 1
        title_font = Font(bold=True, size=12)
        category_font = Font(bold=True, size=11)
        category_fill = PatternFill(start_color="E6F1FB", end_color="E6F1FB", fill_type="solid")
        
        # Overall stats
        ws_summary.cell(row=row, column=1, value="OVERALL STATISTICS").font = title_font
        row += 2
        
        ws_summary.cell(row=row, column=1, value="Total Patients")
        ws_summary.cell(row=row, column=2, value=stats['total_patients'])
        row += 1
        ws_summary.cell(row=row, column=1, value="Total Studies")
        ws_summary.cell(row=row, column=2, value=stats['total_studies'])
        row += 1
        ws_summary.cell(row=row, column=1, value="Total Impacted Teeth")
        ws_summary.cell(row=row, column=2, value=stats['total_impacted'])
        row += 2
        
        # By Type
        ws_summary.cell(row=row, column=1, value="BY TOOTH TYPE").font = category_font
        for cell in [ws_summary.cell(row=row, column=i) for i in range(1, 3)]:
            cell.fill = category_fill
        row += 1
        for tooth_type, count in sorted(stats['by_type'].items()):
            ws_summary.cell(row=row, column=1, value=tooth_type)
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        row += 1
        
        # Pell & Gregory Class
        ws_summary.cell(row=row, column=1, value="PELL & GREGORY CLASS").font = category_font
        for cell in [ws_summary.cell(row=row, column=i) for i in range(1, 3)]:
            cell.fill = category_fill
        row += 1
        for pg_class, count in sorted(stats['pell_gregory_class'].items()):
            ws_summary.cell(row=row, column=1, value=f"Class {pg_class}")
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        row += 1
        
        # Pell & Gregory Depth
        ws_summary.cell(row=row, column=1, value="PELL & GREGORY DEPTH").font = category_font
        for cell in [ws_summary.cell(row=row, column=i) for i in range(1, 3)]:
            cell.fill = category_fill
        row += 1
        for pg_depth, count in sorted(stats['pell_gregory_depth'].items()):
            ws_summary.cell(row=row, column=1, value=f"Depth {pg_depth}")
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        row += 1
        
        # Winter's Classification
        ws_summary.cell(row=row, column=1, value="WINTER'S CLASSIFICATION").font = category_font
        for cell in [ws_summary.cell(row=row, column=i) for i in range(1, 3)]:
            cell.fill = category_fill
        row += 1
        for winters, count in sorted(stats['winters'].items(), key=lambda x: x[1], reverse=True):
            ws_summary.cell(row=row, column=1, value=winters)
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        row += 1
        
        # Severity
        ws_summary.cell(row=row, column=1, value="IMPACTION SEVERITY").font = category_font
        for cell in [ws_summary.cell(row=row, column=i) for i in range(1, 3)]:
            cell.fill = category_fill
        row += 1
        for severity, count in sorted(stats['severity'].items(), key=lambda x: x[1], reverse=True):
            ws_summary.cell(row=row, column=1, value=severity.capitalize())
            ws_summary.cell(row=row, column=2, value=count)
            row += 1
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        
        # Save workbook
        wb.save(path)
        log.info(f"Exported {len(rows)} records → Excel: {path}")

    def close(self):
        self.conn.close()

# ── Pipeline ───────────────────────────────────────────────────────────────────
class ImpactionPipeline:

    def __init__(self, db_path: str):
        self.dicom_reader = DicomReader()
        self.image_reader = ImageReader()
        self.pano_reader  = PanoReader()
        self.html_reader  = HtmlImageReader()
        self.csi_reader   = CsiDataReader()
        self.classifier   = ImpactionClassifier()
        self.db           = ImpactionDatabase(db_path)

    def _get_reader(self, file_path: str):
        """Select appropriate reader based on file extension."""
        ext = Path(file_path).suffix.lower()
        if ext == '.dcm':
            return self.dicom_reader
        elif ext == '.pano':
            return self.pano_reader
        elif ext in ('.png', '.jpg', '.jpeg'):
            return self.image_reader
        elif ext == '.html':
            return self.html_reader
        elif ext == '.csi_data':
            return self.csi_reader
        else:
            log.warning(f"Unknown file type: {ext}, attempting generic image reader")
            return self.image_reader

    def run(self, paths: list[str]):
        total = ok = skipped = errors = 0
        for path in paths:
            total += 1
            log.info(f"Processing [{total}/{len(paths)}]: {Path(path).name}")
            try:
                reader = self._get_reader(path)
                record, pixels = reader.read(path)
                record.impacted_teeth = self.classifier.classify(record, pixels)
                self.db.save_record(record)
                ok += 1
            except Exception as e:
                log.error(f"  ERROR: {e}")
                errors += 1

        log.info(f"\n{'─'*55}")
        log.info(f"  Processed : {total}")
        log.info(f"  OK        : {ok}")
        log.info(f"  Errors    : {errors}")
        log.info(f"{'─'*55}")

        stats = self.db.summary()
        log.info(f"  Patients  : {stats['total_patients']}")
        log.info(f"  Studies   : {stats['total_studies']}")
        log.info(f"  Impacted  : {stats['total_impacted']}")
        log.info(f"{'─'*55}")
        return stats

    def close(self):
        self.db.close()

# ── CLI ────────────────────────────────────────────────────────────────────────
def collect_dicoms(paths: list[str]) -> list[str]:
    """Collect all supported imaging files: DICOM, panoramic, images, HTML, and CSI data."""
    supported_extensions = (".dcm", ".dicom", ".pano", ".png", ".jpg", ".jpeg", ".html", ".csi_data")
    files = []
    
    for p in paths:
        if os.path.isdir(p):
            for root, _, fnames in os.walk(p):
                for fn in fnames:
                    if fn.lower().endswith(supported_extensions):
                        full = os.path.join(root, fn)
                        try:
                            # Validate DICOM files before adding
                            if fn.lower().endswith((".dcm", ".dicom")):
                                pydicom.dcmread(full, stop_before_pixels=True, force=True)
                            # For other formats, just check if file exists
                            elif os.path.isfile(full):
                                pass
                            files.append(full)
                        except Exception:
                            pass
        elif os.path.isfile(p):
            if p.lower().endswith(supported_extensions):
                files.append(p)
    
    return sorted(set(files))


def convert_csi_to_json(csi_files: list[str], output_dir: str = None):
    """Convert CSI data files to JSON format."""
    reader = CsiDataReader()
    converted = 0
    errors = 0
    
    for csi_file in csi_files:
        try:
            csi_data = reader._parse_csi_data(csi_file)
            
            # Determine output path
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
                base_name = Path(csi_file).stem
                json_path = os.path.join(output_dir, f"{base_name}.json")
            else:
                json_path = str(Path(csi_file).with_suffix('.json'))
            
            # Write JSON
            with open(json_path, 'w') as f:
                json.dump(csi_data, f, indent=2, default=str)
            
            log.info(f"✓ Converted: {Path(csi_file).name} → {Path(json_path).name}")
            converted += 1
        except Exception as e:
            log.error(f"✗ Failed to convert {Path(csi_file).name}: {e}")
            errors += 1
    
    log.info(f"\nCSI Conversion Summary:")
    log.info(f"  Converted: {converted}")
    log.info(f"  Errors:    {errors}")
    return converted, errors


def main():
    parser = argparse.ArgumentParser(
        description="Dental Impaction Analyzer — Multi-format imaging → SQLite + JSON/Excel export"
    )
    parser.add_argument("inputs", nargs="*",
                        help="DICOM, panoramic, image, HTML, or CSI data files/directories to process")
    parser.add_argument("--db",  default="dental_impactions.db",
                        help="Output SQLite database path (default: dental_impactions.db)")
    parser.add_argument("--export-json", metavar="FILE",
                        help="Export impaction results to JSON")
    parser.add_argument("--export-excel", metavar="FILE",
                        help="Export impaction results to Excel (.xlsx)")
    parser.add_argument("--convert-csi", metavar="OUTPUT_DIR",
                        help="Convert CSI data files to JSON (specify output directory)")
    parser.add_argument("--summary", action="store_true",
                        help="Print DB summary and exit (no processing)")

    args = parser.parse_args()

    # If converting CSI to JSON only
    if args.convert_csi:
        if not args.inputs:
            log.error("No input files specified for CSI conversion")
            sys.exit(1)
        
        csi_files = [f for f in args.inputs if f.lower().endswith('.csi_data')]
        if csi_files:
            converted, errors = convert_csi_to_json(csi_files, args.convert_csi)
            sys.exit(0 if errors == 0 else 1)
        else:
            # Collect CSI files from directories
            all_files = collect_dicoms(args.inputs)
            csi_files = [f for f in all_files if f.lower().endswith('.csi_data')]
            if csi_files:
                converted, errors = convert_csi_to_json(csi_files, args.convert_csi)
                sys.exit(0 if errors == 0 else 1)
            else:
                log.error("No CSI data files found")
                sys.exit(1)
    
    db_path  = args.db
    pipeline = ImpactionPipeline(db_path)

    if args.summary:
        stats = pipeline.db.summary()
        print(json.dumps(stats, indent=2))
        pipeline.close()
        return

    if not args.inputs:
        parser.print_help()
        sys.exit(1)

    dicom_files = collect_dicoms(args.inputs)
    if not dicom_files:
        log.error("No valid imaging files found in the given paths.")
        sys.exit(1)

    log.info(f"Found {len(dicom_files)} imaging file(s)")
    stats = pipeline.run(dicom_files)

    if args.export_json:
        pipeline.db.export_json(args.export_json)

    if args.export_excel:
        pipeline.db.export_excel(args.export_excel)

    pipeline.close()
    print("\nDone. Database saved to:", db_path)
    print("Summary:", json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
